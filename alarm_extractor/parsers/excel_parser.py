"""
excel_parser.py

Excel parser implementation.

Responsibilities
----------------

✓ Read Excel workbooks
✓ Read every worksheet
✓ Convert worksheet into markdown
✓ Convert workbook into Document model

Does NOT

✗ Save JSON
✗ Call LLM
✗ Generate structured alarms
✗ Write output files
"""

from __future__ import annotations

from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

from models.document_model import (
    Document,
    Metadata,
    Page,
)

from parsers.base_parser import BaseParser


class ExcelParser(BaseParser):
    """
    Excel parser implementation.

    Every worksheet inside a workbook becomes one
    Page inside the universal Document model.
    """

    def __init__(self):

        super().__init__()

        self.console = Console()

    # ---------------------------------------------------------

    def load_excel(
        self,
        excel_file: str | Path,
    ):
        """
        Load an Excel workbook.

        Parameters
        ----------
        excel_file : Path

        Returns
        -------
        openpyxl Workbook
        """

        excel_file = Path(excel_file)

        self.console.print(
            f"[cyan]Loading Workbook[/cyan] {excel_file.name}"
        )

        workbook = load_workbook(
            filename=excel_file,
            data_only=True,
        )

        return workbook

    # ---------------------------------------------------------

    def clean_dataframe(
        self,
        dataframe: pd.DataFrame,
    ) -> pd.DataFrame:
        """
        Remove empty rows and columns.
        """

        dataframe = dataframe.dropna(
            axis=0,
            how="all",
        )

        dataframe = dataframe.dropna(
            axis=1,
            how="all",
        )

        dataframe = dataframe.fillna("")

        return dataframe

    # ---------------------------------------------------------

    def worksheet_to_dataframe(
        self,
        worksheet: Worksheet,
    ) -> pd.DataFrame:
        """
        Convert an Excel worksheet into a DataFrame.
        """

        rows = list(
            worksheet.values
        )

        if not rows:

            return pd.DataFrame()

        headers = rows[0]

        data = rows[1:]

        dataframe = pd.DataFrame(
            data,
            columns=headers,
        )

        dataframe = self.clean_dataframe(
            dataframe
        )

        return dataframe

    # ---------------------------------------------------------

    def dataframe_to_markdown(
        self,
        dataframe: pd.DataFrame,
    ) -> str:
        """
        Convert a dataframe into markdown.
        """

        if dataframe.empty:

            return ""

        try:

            return dataframe.to_markdown(
                index=False
            )

        except Exception:

            return dataframe.to_string(
                index=False
            )
    
        # ---------------------------------------------------------

    def worksheet_to_text(
        self,
        worksheet: Worksheet,
    ) -> str:
        """
        Convert an Excel worksheet into a markdown document.

        Each worksheet is represented as:

        # Sheet: SheetName

        | Table |
        |-------|
        """

        dataframe = self.worksheet_to_dataframe(
            worksheet
        )

        markdown = self.dataframe_to_markdown(
            dataframe
        )

        output = []

        output.append(
            f"# Sheet: {worksheet.title}"
        )

        output.append("")

        if markdown:

            output.append(markdown)

        else:

            output.append(
                "*No data found in worksheet.*"
            )

        return "\n".join(output)

    # ---------------------------------------------------------

    def worksheet_to_page(
        self,
        worksheet: Worksheet,
        page_number: int,
    ) -> Page:
        """
        Convert a worksheet into a Page object.
        """

        self.console.print(
            f"[green]Processing Worksheet[/green] "
            f"{worksheet.title}"
        )

        markdown = self.worksheet_to_text(
            worksheet
        )

        return Page(

            page_number=page_number,

            text=markdown,

            markdown=markdown,

        )

    # ---------------------------------------------------------

    def workbook_to_pages(
        self,
        workbook,
    ) -> List[Page]:
        """
        Convert every worksheet into a Page.

        Workbook

            Sheet1

            Sheet2

            Sheet3

        becomes

            Page1

            Page2

            Page3
        """

        pages = []

        page_number = 1

        for worksheet in workbook.worksheets:

            page = self.worksheet_to_page(

                worksheet,

                page_number,

            )

            pages.append(page)

            page_number += 1

        return pages

    # ---------------------------------------------------------

    def extract_metadata(
        self,
        workbook,
        input_file: Path,
        pages: List[Page],
    ) -> Metadata:
        """
        Create document metadata.
        """

        metadata = Metadata(

            source_file=input_file.name,

            document_type="excel",

            parser="OpenPyXL + Pandas",

            page_count=len(pages),

            extra={

                "sheet_count": len(workbook.sheetnames),

                "sheet_names": workbook.sheetnames,

            },

        )

        return metadata

        # ---------------------------------------------------------

    def parse(
        self,
        input_file: str | Path,
    ) -> Document:
        """
        Parse an Excel workbook and return a Document object.

        Workflow

        Excel
            ↓
        Workbook
            ↓
        Worksheets
            ↓
        Pages
            ↓
        Document
        """

        input_file = Path(input_file)

        self.console.rule(
            "[bold cyan]Excel Parser[/bold cyan]"
        )

        try:

            workbook = self.load_excel(
                input_file
            )

            pages = self.workbook_to_pages(
                workbook
            )

            metadata = self.extract_metadata(
                workbook,
                input_file,
                pages,
            )

            document = Document(

                metadata=metadata,

                pages=pages,

            )

            self.console.print(

                f"[bold green]Successfully Parsed[/bold green] "

                f"{metadata.page_count} worksheet(s)"

            )

            return document

        except Exception as e:

            self.console.print(
                f"[bold red]Excel Parsing Failed[/bold red]"
            )

            self.console.print(
                f"[red]{str(e)}[/red]"
            )

            raise